# truck_tracking_web_service_complete.py - Servicio completo para web
import pandas as pd
from datetime import datetime, timedelta
import logging
import threading
import os
import pymysql
import requests
from shapely.geometry import Point, Polygon
import json
from typing import List, Dict, Tuple, Optional
import time

logger = logging.getLogger(__name__)


class TruckTrackingWebServiceComplete:
    """
    Servicio web completo que integra TODA la funcionalidad del sistema original
    """

    def __init__(self, config):
        """Inicializa el servicio web completo"""
        self.config = config
        self.source_connection = None
        self.target_connection = None
        self.last_processing_time = None
        self.processing_lock = threading.Lock()

        # DATOS Y CACHE
        self.geocercas = {}
        self.historical_data = {}
        self.results_data = []
        self.cache = {
            'trucks_data': [],
            'alerts': {},
            'stats': {},
            'last_update': None
        }

        # Configuración igual que original
        self.geocerca_hierarchy = ['DOCKS', 'TRACK AND TRACE', 'CBN', 'CIUDADES']

        self.alert_config = {
            'critical_hours': 48,
            'warning_hours': 8,
            'normal_hours': 4
        }

        self.deposito_geocerca_mapping = {
            "Cerveceria SCZ": {
                "ciudad": "SANTA CRUZ",
                "cbn": "PLANTA SANTA CRUZ",
                "track_trace": "TYT - PLANTA SANTA CRUZ",
                "docks": "DOCK - 7 - PLANTA SANTA CRUZ"
            },
            "Cerveceria LPZ": {
                "ciudad": "LA PAZ",
                "cbn": "PLANTA LA PAZ",
                "track_trace": "TYT - PLANTA LA PAZ",
                "docks": "DOCK - 3 - PLANTA LA PAZ"
            },
            "Cerveceria CBBA": {
                "ciudad": "COCHABAMBA",
                "cbn": "PLANTA COCHABAMBA",
                "track_trace": "TYT - PLANTA COCHABAMBA",
                "docks": "DOCK - 5 - PLANTA COCHABAMBA"
            }
        }

        # Inicializar sistema
        self._init_system()

    def _init_system(self):
        """Inicializa el sistema completo"""
        try:
            # Conectar bases de datos
            self.connect_databases()

            # Cargar geocercas si existe el archivo
            if os.path.exists(self.config['excel_path']):
                self.load_geocercas()
            else:
                logger.warning(f"Archivo de geocercas no encontrado: {self.config['excel_path']}")

            # Cargar datos históricos si existe el archivo
            if self.config.get('historical_path') and os.path.exists(self.config['historical_path']):
                self.load_historical_data()
            else:
                logger.warning("Archivo histórico no encontrado")

            logger.info("Sistema completo inicializado correctamente")
            return True

        except Exception as e:
            logger.error(f"Error inicializando sistema completo: {e}")
            return False

    def connect_databases(self):
        """Conecta a ambas bases de datos"""
        try:
            # Conexión a BD de origen
            self.source_connection = pymysql.connect(
                cursorclass=pymysql.cursors.DictCursor,
                **self.config['source_db']
            )
            logger.info("Conexión exitosa a BD de origen")

            # Conexión a BD de destino
            self.target_connection = pymysql.connect(
                cursorclass=pymysql.cursors.DictCursor,
                **self.config['target_db']
            )
            logger.info("Conexión exitosa a BD de destino")

            # Crear tabla de tracking si no existe
            self._create_tracking_table()

            # AGREGAR ESTA LÍNEA:
            self._update_table_structure()

            return True

        except Exception as e:
            logger.error(f"Error conectando a bases de datos: {e}")
            return False

    def _create_tracking_table(self):
        """Crea tabla básica (simplificada para que funcione)"""
        try:
            with self.target_connection.cursor() as cursor:
                create_table_query = """
                CREATE TABLE IF NOT EXISTS truck_tracking (
                    id BIGINT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
                    patente VARCHAR(255) NOT NULL,
                    planilla VARCHAR(255) NULL,
                    status VARCHAR(50) NULL,
                    deposito_destino VARCHAR(255) NULL,
                    latitude DECIMAL(10, 8) NULL,
                    longitude DECIMAL(11, 8) NULL,
                    velocidad_kmh DECIMAL(5, 2) NULL,
                    porcentaje_entrega DECIMAL(5, 2) DEFAULT 0.00,
                    estado_entrega VARCHAR(50) DEFAULT 'EN_TRANSITO',
                    tiempo_espera_minutos INT DEFAULT 0,
                    alert_level VARCHAR(20) DEFAULT 'NORMAL',
                    ultima_actualizacion DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    INDEX idx_patente (patente)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
                """
                cursor.execute(create_table_query)
                self.target_connection.commit()
                logger.info("Tabla truck_tracking verificada/creada")
        except Exception as e:
            logger.error(f"Error creando tabla: {e}")

    def get_last_update_from_db(self):
        """Obtiene último update de BD"""
        try:
            with self.source_connection.cursor() as cursor:
                query = "SELECT MAX(created_at) as ultimo_update FROM trucks WHERE status = 'SALIDA'"
                cursor.execute(query)
                result = cursor.fetchone()

                if result and result['ultimo_update']:
                    ultimo_update = result['ultimo_update']
                    if isinstance(ultimo_update, str):
                        ultimo_update = datetime.strptime(ultimo_update, '%Y-%m-%d %H:%M:%S')
                    return ultimo_update.strftime('%H:%M:%S')
                else:
                    return "N/A"
        except Exception as e:
            logger.error(f"Error obteniendo último update: {e}")
            return "Error"

    def get_all_trucks_status_complete(self):
        """Obtiene estado de camiones (versión simplificada por ahora)"""
        try:
            # Por ahora devolver datos básicos para que funcione
            return [
                {
                    'patente': 'TEST-001',
                    'planilla': 'PL001',
                    'status': 'SALIDA',
                    'deposito_destino': 'Cerveceria SCZ',
                    'latitude': -17.783,
                    'longitude': -63.182,
                    'velocidad_kmh': 45.5,
                    'en_docks': 'NO',
                    'en_track_trace': 'NO',
                    'en_cbn': 'NO',
                    'en_ciudades': 'SI en SANTA CRUZ',
                    'porcentaje_entrega': 25.0,
                    'estado_entrega': 'EN_CIUDAD',
                    'tiempo_espera_minutos': 0,
                    'tiempo_espera_horas': 0.0,
                    'alert_level': 'NORMAL'
                }
            ]
        except Exception as e:
            logger.error(f"Error obteniendo camiones: {e}")
            return []

    def get_alerts_summary_complete(self):
        """Obtiene resumen de alertas (simplificado)"""
        return {
            'total_waiting': 0,
            'critical_count': 0,
            'warning_count': 0,
            'attention_count': 0
        }

    def get_geocercas_distribution(self):
        """Obtiene distribución de geocercas (simplificado)"""
        return {
            'docks': 0,
            'track_and_trace': 0,
            'cbn': 0,
            'ciudades': 1
        }

    def get_dashboard_stats_complete(self):
        """Obtiene estadísticas para dashboard (simplificado)"""
        return {
            'total_camiones': 1,
            'en_transito': 1,
            'en_descarga': 0,
            'alertas_criticas': 0,
            'alertas_warning': 0,
            'promedio_progreso': 25.0
        }

    def get_active_alerts_complete(self):
        """Obtiene alertas activas (simplificado)"""
        return []

    def get_critical_alerts_complete(self):
        """Obtiene alertas críticas (simplificado)"""
        return []

    def get_geocercas_status(self):
        """Obtiene estado de geocercas (simplificado)"""
        return []

    def generate_excel_report_complete(self):
        """Genera reporte Excel (simplificado)"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'tracking_simple_{timestamp}.xlsx'

            # Crear DataFrame simple
            data = self.get_all_trucks_status_complete()
            df = pd.DataFrame(data)

            # Guardar Excel básico
            df.to_excel(filename, index=False)
            return filename
        except Exception as e:
            logger.error(f"Error generando Excel: {e}")
            return None

    def process_all_trucks_complete(self):
        """Procesa camiones (simplificado)"""
        logger.info("Procesamiento completo ejecutado (versión simplificada)")
        return True

    # Funciones placeholder para las otras que faltan
    def load_geocercas(self):
        """Carga geocercas (placeholder)"""
        logger.info("Función load_geocercas ejecutada")
        return True

    def load_historical_data(self):
        """Carga datos históricos (placeholder)"""
        logger.info("Función load_historical_data ejecutada")
        return True

    def get_trucks_in_transit(self) -> List[Dict]:
        """Obtiene camiones en tránsito con validación de último viaje"""
        try:
            with self.source_connection.cursor() as cursor:
                query = """
                SELECT t1.cod, t1.deposito_origen, t1.cod_destino, t1.deposito_destino, 
                       t1.planilla, t1.patente, t1.fecha_salida, t1.hora_salida, 
                       t1.fecha_llegada, t1.hora_llegada, t1.cod_producto, t1.producto,
                       t1.status, t1.salida
                FROM trucks t1
                WHERE t1.status = 'SALIDA'
                  AND NOT EXISTS (
                    SELECT 1 
                    FROM trucks t2 
                    WHERE t2.patente = t1.patente 
                      AND (
                        t2.fecha_salida > t1.fecha_salida 
                        OR (t2.fecha_salida = t1.fecha_salida AND t2.hora_salida > t1.hora_salida)
                      )
                  )
                ORDER BY t1.fecha_salida DESC, t1.hora_salida DESC
                """

                cursor.execute(query)
                trucks = cursor.fetchall()
                logger.info(f"Encontrados {len(trucks)} camiones en tránsito")
                return trucks

        except Exception as e:
            logger.error(f"Error obteniendo camiones: {e}")
            return []

    def get_all_trucks_locations_parallel(self, trucks: List[Dict]) -> Dict[str, Dict]:
        """Obtiene ubicaciones de múltiples camiones en paralelo desde API"""
        try:
            headers = {
                'token': self.config['api']['token'],
                'Content-Type': 'application/json'
            }

            response = requests.get(
                f"{self.config['api']['base_url']}/ultimaubicaciontodos",
                headers=headers,
                timeout=30
            )

            if response.status_code == 200:
                data = response.json()
                locations = {}

                for vehicle in data:
                    patente = vehicle.get('id_unidad')
                    if patente:
                        locations[patente] = {
                            'patente': patente,
                            'latitude': vehicle.get('latitud'),
                            'longitude': vehicle.get('longitud'),
                            'timestamp': vehicle.get('tiempoMovimientoFormatted'),
                            'speed': vehicle.get('velocidad_kmh', 0),
                            'direction': vehicle.get('direccion', 0)
                        }

                logger.info(f"API devolvió ubicaciones para {len(locations)} vehículos")
                return locations
            else:
                logger.error(f"Error API: Status {response.status_code}")
                return {}

        except Exception as e:
            logger.error(f"Error obteniendo ubicaciones: {e}")
            return {}

    def load_geocercas(self):
        """Carga las geocercas desde el archivo Excel"""
        try:
            df = pd.read_excel(self.config['excel_path'])
            logger.info(f"Excel de geocercas cargado: {len(df)} filas")

            # Buscar columnas
            col_nombregrupo = None
            col_nombregeocerca = None
            col_puntos = None

            for col in df.columns:
                if 'NOMBREGRUPO' in col.upper():
                    col_nombregrupo = col
                    logger.info(f"NOMBREGRUPO encontrado: '{col}'")
                if 'NOMBREGEOCERCA' in col.upper():
                    col_nombregeocerca = col
                    logger.info(f"NOMBREGEOCERCA encontrado: '{col}'")
                if 'PUNTOS' in col.upper() and 'GEOCERCA' in col.upper():
                    col_puntos = col
                    logger.info(f"PUNTOS GEOCERCA encontrado: '{col}'")

            if not all([col_nombregrupo, col_nombregeocerca, col_puntos]):
                logger.error("Columnas requeridas no encontradas en Excel")
                return False

            # Procesar geocercas
            procesadas = 0
            errores = 0

            for idx, row in df.iterrows():
                try:
                    grupo = str(row[col_nombregrupo]).strip()
                    nombre_geocerca = str(row[col_nombregeocerca]).strip()
                    puntos_str = str(row[col_puntos]).strip()

                    if all([grupo != 'nan', nombre_geocerca != 'nan', puntos_str != 'nan']):
                        puntos = self._parse_geocerca_points(puntos_str)

                        if grupo not in self.geocercas:
                            self.geocercas[grupo] = []

                        self.geocercas[grupo].append({
                            'nombre': nombre_geocerca,
                            'puntos': puntos,
                            'polygon': Polygon(puntos) if len(puntos) >= 3 else None
                        })
                        procesadas += 1

                except Exception as e:
                    logger.warning(f"Error procesando geocerca fila {idx}: {e}")
                    errores += 1

            logger.info(f"Geocercas procesadas: {procesadas}, errores: {errores}")
            for grupo, geocercas_lista in self.geocercas.items():
                validas = sum(1 for g in geocercas_lista if g['polygon'] is not None)
                logger.info(f"  {grupo}: {len(geocercas_lista)} geocercas ({validas} válidas)")

            return True

        except Exception as e:
            logger.error(f"Error cargando geocercas: {e}")
            return False

    def _parse_geocerca_points(self, puntos_str: str) -> List[Tuple[float, float]]:
        """Parsea los puntos de coordenadas de una geocerca"""
        try:
            puntos = []
            coordenadas = puntos_str.split(',')

            for coord in coordenadas:
                coord = coord.strip()
                if coord and ' ' in coord:
                    parts = coord.split()
                    if len(parts) >= 2:
                        try:
                            lat = float(parts[0])
                            lng = float(parts[1])
                            puntos.append((lng, lat))  # Shapely usa (lng, lat)
                        except ValueError:
                            continue

            return puntos
        except Exception as e:
            logger.warning(f"Error parseando puntos: {e}")
            return []

    def check_point_in_geocercas(self, lat: float, lng: float, deposito_destino: str = None) -> Dict[str, str]:
        """Verifica en qué geocercas se encuentra un punto con mapeo correlativo"""
        point = Point(lng, lat)
        result = {
            'DOCKS': 'NO',
            'TRACK AND TRACE': 'NO',
            'CBN': 'NO',
            'CIUDADES': 'NO'
        }

        # Mapeo específico para el depósito
        target_geocercas = {}
        if deposito_destino and deposito_destino in self.deposito_geocerca_mapping:
            mapping = self.deposito_geocerca_mapping[deposito_destino]
            target_geocercas = {
                'CIUDADES': mapping.get('ciudad'),
                'CBN': mapping.get('cbn'),
                'TRACK AND TRACE': mapping.get('track_trace'),
                'DOCKS': mapping.get('docks')
            }

        # Verificar en orden jerárquico
        for grupo in self.geocerca_hierarchy:
            if grupo in self.geocercas:
                target_name = target_geocercas.get(grupo)

                if target_name:
                    # Buscar geocerca específica
                    for geocerca in self.geocercas[grupo]:
                        if (target_name.upper() in geocerca['nombre'].upper() or
                                geocerca['nombre'].upper() in target_name.upper()):
                            if geocerca['polygon'] and geocerca['polygon'].contains(point):
                                result[grupo] = f"SI en {geocerca['nombre']}"
                                break

                # Si no encontramos la específica, buscar cualquiera
                if result[grupo] == 'NO':
                    for geocerca in self.geocercas[grupo]:
                        if geocerca['polygon'] and geocerca['polygon'].contains(point):
                            result[grupo] = f"SI en {geocerca['nombre']}"
                            break

        return result

    def calculate_delivery_progress(self, geocerca_status: Dict[str, str], deposito_destino: str = None) -> Tuple[float, str]:
        """Calcula porcentaje de progreso de entrega basado en geocercas"""
        porcentaje = 0.0
        estado = "EN_TRANSITO"

        # Sistema de progreso jerárquico
        if geocerca_status['CIUDADES'] != 'NO':
            porcentaje += 25.0
            estado = "EN_CIUDAD"

        if geocerca_status['CBN'] != 'NO':
            porcentaje += 25.0
            estado = "EN_CENTRO_DISTRIBUCION"

        if geocerca_status['TRACK AND TRACE'] != 'NO':
            porcentaje += 30.0
            estado = "EN_ZONA_DESCARGA"

        if geocerca_status['DOCKS'] != 'NO':
            porcentaje += 20.0
            estado = "DESCARGANDO"

        # Si está en los 4 puntos = 100%
        all_present = all(status != 'NO' for status in geocerca_status.values())
        if all_present:
            porcentaje = 100.0
            estado = "DESCARGANDO_CONFIRMADO"

        return min(porcentaje, 100.0), estado

    def calculate_waiting_time_for_discharge(self, truck_data: Dict, geocerca_status: Dict[str, str], current_estado_entrega: str) -> Tuple[int, str, str, str]:
        """Calcula tiempo de espera para descarga con datos históricos"""
        try:
            patente = truck_data['patente']
            planilla = truck_data['planilla']

            # Verificar si está esperando descarga
            is_waiting_for_discharge = truck_data.get('status', '') != 'SALIDA'

            if not is_waiting_for_discharge:
                is_waiting_for_discharge = (
                        geocerca_status['DOCKS'] != 'NO' or
                        geocerca_status['TRACK AND TRACE'] != 'NO' or
                        current_estado_entrega in ['EN_ZONA_DESCARGA', 'DESCARGANDO', 'DESCARGANDO_CONFIRMADO']
                )

            # Determinar estado específico de descarga
            if truck_data.get('status', '') != 'SALIDA':
                estado_descarga = f"STATUS_{truck_data.get('status', 'UNKNOWN')}"
            elif geocerca_status['DOCKS'] != 'NO':
                estado_descarga = 'EN_DOCKS'
            elif geocerca_status['TRACK AND TRACE'] != 'NO':
                estado_descarga = 'EN_TRACK_TRACE'
            elif current_estado_entrega in ['DESCARGANDO', 'DESCARGANDO_CONFIRMADO']:
                estado_descarga = 'DESCARGANDO'
            elif current_estado_entrega == 'EN_ZONA_DESCARGA':
                estado_descarga = 'ZONA_DESCARGA'
            else:
                estado_descarga = 'NO_ESPERANDO'

            if not is_waiting_for_discharge:
                return 0, None, 'NO_ESPERANDO', 'NORMAL'

            inicio_espera = None

            # Buscar en datos históricos del Excel
            if patente in self.historical_data:
                inicio_espera = self.historical_data[patente]['primera_entrada_descarga']

            # Si no hay datos históricos, buscar en BD
            if inicio_espera is None:
                with self.target_connection.cursor() as cursor:
                    history_query = """
                    SELECT primera_deteccion
                    FROM truck_tracking 
                    WHERE patente = %s AND planilla = %s
                    AND (estado_entrega IN ('EN_ZONA_DESCARGA', 'DESCARGANDO', 'DESCARGANDO_CONFIRMADO'))
                    ORDER BY primera_deteccion ASC
                    LIMIT 1
                    """

                    cursor.execute(history_query, (patente, planilla))
                    result = cursor.fetchone()

                    if result:
                        inicio_espera = result['primera_deteccion']

            # Si no hay registro, este es el primero
            if inicio_espera is None:
                inicio_espera = datetime.now()

            # Calcular tiempo transcurrido
            if isinstance(inicio_espera, str):
                inicio_espera = datetime.strptime(inicio_espera, '%Y-%m-%d %H:%M:%S')

            tiempo_espera = datetime.now() - inicio_espera
            tiempo_espera_minutos = int(tiempo_espera.total_seconds() / 60)

            # Determinar nivel de alerta
            tiempo_espera_horas = tiempo_espera_minutos / 60
            if tiempo_espera_horas >= self.alert_config['critical_hours']:
                alert_level = 'CRITICAL'
            elif tiempo_espera_horas >= self.alert_config['warning_hours']:
                alert_level = 'WARNING'
            elif tiempo_espera_horas >= self.alert_config['normal_hours']:
                alert_level = 'ATTENTION'
            else:
                alert_level = 'NORMAL'

            return tiempo_espera_minutos, inicio_espera.strftime('%Y-%m-%d %H:%M:%S'), estado_descarga, alert_level

        except Exception as e:
            logger.error(f"Error calculando tiempo de espera para {truck_data['patente']}: {e}")
            return 0, None, 'ERROR', 'ERROR'

    def _adjust_time_utc_minus_4(self, time_input) -> str:
        """Ajusta la hora restando 1 hora (UTC-4)"""
        try:
            if not time_input:
                return ""

            if isinstance(time_input, str):
                try:
                    time_obj = datetime.strptime(time_input, '%H:%M:%S').time()
                except:
                    return time_input
            elif hasattr(time_input, 'hour'):
                time_obj = time_input
            elif hasattr(time_input, 'time'):
                time_obj = time_input.time()
            else:
                return str(time_input)

            dt = datetime.combine(datetime.today(), time_obj)
            adjusted_dt = dt - timedelta(hours=1)

            if adjusted_dt.date() < datetime.today().date():
                adjusted_dt = adjusted_dt + timedelta(days=1)

            return adjusted_dt.time().strftime('%H:%M:%S')

        except Exception as e:
            logger.warning(f"Error ajustando hora UTC-4: {e}")
            return str(time_input) if time_input else ""

    def load_historical_data(self):
        """Carga datos históricos del Excel DataGrid"""
        try:
            if not os.path.exists(self.config['historical_path']):
                logger.warning("Archivo histórico no encontrado")
                return True

            df_history = pd.read_excel(self.config['historical_path'])
            logger.info(f"Archivo histórico cargado: {len(df_history)} registros")

            if len(df_history) == 0:
                return True

            # Mapear columnas (flexible para diferentes nombres)
            col_mapping = {}
            for col in df_history.columns:
                col_upper = col.upper()
                if 'PLACA' in col_upper:
                    col_mapping['patente'] = col
                elif 'ENTRADA' in col_upper and 'FECHA' in col_upper:
                    col_mapping['fecha_entrada'] = col
                elif 'GEOCERCA' in col_upper and 'NOMBRE' in col_upper:
                    col_mapping['geocerca'] = col
                elif 'GRUPO' in col_upper and 'GEOCERCA' in col_upper:
                    col_mapping['grupo_geocerca'] = col

            # Procesar datos históricos
            processed_trucks = {}
            for idx, row in df_history.iterrows():
                try:
                    patente = str(row.get(col_mapping.get('patente', ''), '')).strip()
                    geocerca = str(row.get(col_mapping.get('geocerca', ''), '')).strip()
                    grupo = str(row.get(col_mapping.get('grupo_geocerca', ''), '')).strip()
                    fecha_entrada_str = str(row.get(col_mapping.get('fecha_entrada', ''), '')).strip()

                    if not patente or patente == 'nan':
                        continue

                    # Verificar si es geocerca de descarga
                    is_descarga_geocerca = (
                            'DOCK' in geocerca.upper() or
                            'TRACK' in geocerca.upper() or
                            'TYT' in geocerca.upper() or
                            ('PLANTA' in geocerca.upper() and grupo.upper() == 'CBN')
                    )

                    if is_descarga_geocerca and fecha_entrada_str and fecha_entrada_str != 'nan':
                        try:
                            fecha_entrada = pd.to_datetime(fecha_entrada_str, format='%d/%m/%Y %H:%M:%S')

                            if patente not in processed_trucks:
                                processed_trucks[patente] = {
                                    'primera_entrada_descarga': fecha_entrada,
                                    'geocerca_inicial': geocerca,
                                    'grupo_inicial': grupo
                                }
                            elif fecha_entrada < processed_trucks[patente]['primera_entrada_descarga']:
                                processed_trucks[patente]['primera_entrada_descarga'] = fecha_entrada
                                processed_trucks[patente]['geocerca_inicial'] = geocerca

                        except Exception as e:
                            logger.warning(f"Error parseando fecha {fecha_entrada_str}: {e}")

                except Exception as e:
                    logger.warning(f"Error procesando fila histórica {idx}: {e}")

            self.historical_data = processed_trucks
            logger.info(f"Datos históricos procesados para {len(processed_trucks)} camiones")
            return True

        except Exception as e:
            logger.error(f"Error cargando datos históricos: {e}")
            return True

    def get_all_trucks_status_complete(self):
        """Obtiene estado completo de todos los camiones con TODAS las funcionalidades"""
        try:
            # Si tenemos cache reciente (< 5 minutos), usarlo
            if (self.cache['last_update'] and
                    (datetime.now() - self.cache['last_update']).seconds < 300):
                return self.cache['trucks_data']

            # Obtener datos frescos
            trucks = self.get_trucks_in_transit()
            if not trucks:
                return []

            # Obtener ubicaciones
            locations = self.get_all_trucks_locations_parallel(trucks)

            # Procesar datos completos
            trucks_data = []
            for truck in trucks:
                patente = truck['patente']
                location = locations.get(patente)

                if location and location['latitude'] and location['longitude']:
                    # Verificar geocercas con mapeo correlativo
                    geocerca_status = self.check_point_in_geocercas(
                        location['latitude'],
                        location['longitude'],
                        truck.get('deposito_destino', '')
                    )

                    # Calcular progreso de entrega
                    porcentaje_entrega, estado_entrega = self.calculate_delivery_progress(
                        geocerca_status, truck.get('deposito_destino', '')
                    )

                    # Calcular tiempo de espera
                    tiempo_espera_minutos, inicio_espera_str, estado_descarga, alert_level = \
                        self.calculate_waiting_time_for_discharge(
                            truck, geocerca_status, estado_entrega
                        )

                    truck_data = {
                        'patente': patente,
                        'planilla': truck.get('planilla', ''),
                        'status': truck.get('status', ''),
                        'deposito_origen': truck.get('deposito_origen', ''),
                        'deposito_destino': truck.get('deposito_destino', ''),
                        'producto': truck.get('producto', ''),
                        'cod_producto': truck.get('cod_producto', ''),
                        'salida': truck.get('salida', 0),
                        'fecha_salida': str(truck.get('fecha_salida', '')),
                        'hora_salida': self._adjust_time_utc_minus_4(truck.get('hora_salida', '')),
                        'fecha_llegada': str(truck.get('fecha_llegada', '')),
                        'hora_llegada': self._adjust_time_utc_minus_4(truck.get('hora_llegada', '')),
                        'latitude': location.get('latitude'),
                        'longitude': location.get('longitude'),
                        'velocidad_kmh': location.get('speed', 0),
                        'direccion': location.get('direction', 0),
                        'timestamp': location.get('timestamp', ''),
                        'en_docks': geocerca_status['DOCKS'],
                        'en_track_trace': geocerca_status['TRACK AND TRACE'],
                        'en_cbn': geocerca_status['CBN'],
                        'en_ciudades': geocerca_status['CIUDADES'],
                        'porcentaje_entrega': porcentaje_entrega,
                        'estado_entrega': estado_entrega,
                        'tiempo_espera_minutos': tiempo_espera_minutos,
                        'tiempo_espera_horas': round(tiempo_espera_minutos / 60, 2) if tiempo_espera_minutos > 0 else 0,
                        'estado_descarga': estado_descarga,
                        'alert_level': alert_level,
                        'inicio_espera': inicio_espera_str,
                        'fecha_proceso': datetime.now().isoformat()
                    }

                    trucks_data.append(truck_data)

                    # Guardar en BD de destino
                    self._save_truck_tracking_complete(truck, location, geocerca_status,
                                                       porcentaje_entrega, estado_entrega,
                                                       tiempo_espera_minutos, estado_descarga, inicio_espera_str)

            # Actualizar cache
            self.cache['trucks_data'] = trucks_data
            self.cache['last_update'] = datetime.now()

            return trucks_data

        except Exception as e:
            logger.error(f"Error obteniendo estado completo de camiones: {e}")
            return []

    def _save_truck_tracking_complete(self, truck_data: Dict, location_data: Dict, geocerca_status: Dict[str, str],
                                      porcentaje_entrega: float, estado_entrega: str, tiempo_espera_minutos: int,
                                      estado_descarga: str, inicio_espera_str: str):
        """Guarda tracking completo en BD de destino"""
        try:
            with self.target_connection.cursor() as cursor:
                # Verificar si ya existe
                check_query = """
                SELECT id, inicio_espera_descarga FROM truck_tracking 
                WHERE patente = %s AND planilla = %s
                """
                cursor.execute(check_query, (truck_data['patente'], truck_data['planilla']))
                existing = cursor.fetchone()

                hora_salida_adjusted = self._adjust_time_utc_minus_4(truck_data.get('hora_salida'))
                hora_llegada_adjusted = self._adjust_time_utc_minus_4(truck_data.get('hora_llegada'))

                if existing:
                    # Actualizar registro existente
                    inicio_espera_final = existing['inicio_espera_descarga'] or inicio_espera_str

                    update_query = """
                    UPDATE truck_tracking SET
                        latitude = %s, longitude = %s, velocidad_kmh = %s, direccion = %s, 
                        geocerca_docks = %s, geocerca_track_trace = %s, geocerca_cbn = %s, geocerca_ciudades = %s,
                        porcentaje_entrega = %s, estado_entrega = %s, tiempo_espera_minutos = %s, estado_descarga = %s,
                        hora_salida = %s, hora_llegada = %s, status = %s, salida = %s,
                        inicio_espera_descarga = %s, alert_level = %s, ultima_actualizacion = NOW()
                    WHERE patente = %s AND planilla = %s
                    """

                    cursor.execute(update_query, (
                        location_data['latitude'], location_data['longitude'], location_data['speed'],
                        location_data['direction'],
                        geocerca_status['DOCKS'], geocerca_status['TRACK AND TRACE'],
                        geocerca_status['CBN'], geocerca_status['CIUDADES'], porcentaje_entrega, estado_entrega,
                        tiempo_espera_minutos, estado_descarga, hora_salida_adjusted, hora_llegada_adjusted,
                        truck_data.get('status'), truck_data.get('salida'), inicio_espera_final,
                        self._get_alert_level(tiempo_espera_minutos),
                        truck_data['patente'], truck_data['planilla']
                    ))
                else:
                    # Insertar nuevo registro
                    insert_query = """
                    INSERT INTO truck_tracking (
                        cod, deposito_origen, cod_destino, deposito_destino, planilla, patente,
                        fecha_salida, hora_salida, fecha_llegada, hora_llegada, cod_producto, producto,
                        latitude, longitude, velocidad_kmh, direccion,
                        geocerca_docks, geocerca_track_trace, geocerca_cbn, geocerca_ciudades,
                        porcentaje_entrega, estado_entrega, status, salida,
                        tiempo_espera_minutos, estado_descarga, inicio_espera_descarga, alert_level
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s
                    )
                    """

                    cursor.execute(insert_query, (
                        truck_data.get('cod'), truck_data.get('deposito_origen'), truck_data.get('cod_destino'),
                        truck_data.get('deposito_destino'), truck_data.get('planilla'), truck_data['patente'],
                        truck_data.get('fecha_salida'), hora_salida_adjusted, truck_data.get('fecha_llegada'),
                        hora_llegada_adjusted, truck_data.get('cod_producto'), truck_data.get('producto'),
                        location_data['latitude'], location_data['longitude'], location_data['speed'],
                        location_data['direction'], geocerca_status['DOCKS'], geocerca_status['TRACK AND TRACE'],
                        geocerca_status['CBN'], geocerca_status['CIUDADES'], porcentaje_entrega, estado_entrega,
                        truck_data.get('status'), truck_data.get('salida'), tiempo_espera_minutos, estado_descarga,
                        inicio_espera_str, self._get_alert_level(tiempo_espera_minutos)
                    ))

                self.target_connection.commit()

        except Exception as e:
            logger.error(f"Error guardando tracking completo de {truck_data['patente']}: {e}")
            self.target_connection.rollback()

    def _get_alert_level(self, tiempo_espera_minutos: int) -> str:
        """Determina nivel de alerta basado en tiempo de espera"""
        horas = tiempo_espera_minutos / 60
        if horas >= self.alert_config['critical_hours']:
            return 'CRITICAL'
        elif horas >= self.alert_config['warning_hours']:
            return 'WARNING'
        elif horas >= self.alert_config['normal_hours']:
            return 'ATTENTION'
        else:
            return 'NORMAL'

    def _create_tracking_table(self):
        """Crea la tabla de tracking completa con todas las columnas"""
        try:
            with self.target_connection.cursor() as cursor:
                create_table_query = """
                CREATE TABLE IF NOT EXISTS truck_tracking (
                    id BIGINT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
                    cod VARCHAR(255) NULL,
                    deposito_origen VARCHAR(255) NULL,
                    cod_destino VARCHAR(255) NULL,
                    deposito_destino VARCHAR(255) NULL,
                    planilla VARCHAR(255) NULL,
                    patente VARCHAR(255) NOT NULL,
                    fecha_salida DATE NULL,
                    hora_salida TIME NULL,
                    fecha_llegada DATE NULL,
                    hora_llegada TIME NULL,
                    cod_producto VARCHAR(255) NULL,
                    producto VARCHAR(255) NULL,
                    status VARCHAR(50) NULL,
                    salida INT NULL,

                    -- Datos de ubicación actual
                    latitude DECIMAL(10, 8) NULL,
                    longitude DECIMAL(11, 8) NULL,
                    velocidad_kmh DECIMAL(5, 2) NULL,
                    direccion INT NULL,

                    -- Estados de geocercas
                    geocerca_docks VARCHAR(255) DEFAULT 'NO',
                    geocerca_track_trace VARCHAR(255) DEFAULT 'NO',
                    geocerca_cbn VARCHAR(255) DEFAULT 'NO',
                    geocerca_ciudades VARCHAR(255) DEFAULT 'NO',

                    -- Porcentaje de progreso de entrega
                    porcentaje_entrega DECIMAL(5, 2) DEFAULT 0.00,
                    estado_entrega VARCHAR(50) DEFAULT 'EN_TRANSITO',

                    -- Tiempo de espera para descarga
                    inicio_espera_descarga DATETIME NULL,
                    tiempo_espera_minutos INT DEFAULT 0,
                    estado_descarga VARCHAR(50) DEFAULT 'NO_INICIADO',
                    alert_level VARCHAR(20) DEFAULT 'NORMAL',

                    -- Control de actualizaciones
                    primera_deteccion DATETIME DEFAULT CURRENT_TIMESTAMP,
                    ultima_actualizacion DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,

                    UNIQUE KEY unique_patente_planilla (patente, planilla),
                    INDEX idx_patente (patente),
                    INDEX idx_status (status),
                    INDEX idx_porcentaje_entrega (porcentaje_entrega),
                    INDEX idx_tiempo_espera (tiempo_espera_minutos),
                    INDEX idx_alert_level (alert_level)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;
                """

                cursor.execute(create_table_query)
                self.target_connection.commit()
                logger.info("Tabla truck_tracking completa verificada/creada")

        except Exception as e:
            logger.error(f"Error creando tabla completa: {e}")

    def generate_waiting_alerts_complete(self):
        """Genera alertas de tiempo de espera (versión simple sin alert_level)"""
        try:
            with self.target_connection.cursor() as cursor:
                # Query simple sin columnas problemáticas
                alerts_query = """
                SELECT patente, planilla, deposito_destino, tiempo_espera_minutos, status,
                       ROUND(tiempo_espera_minutos / 60.0, 1) as horas_espera
                FROM truck_tracking 
                WHERE tiempo_espera_minutos > %s 
                ORDER BY tiempo_espera_minutos DESC
                """

                cursor.execute(alerts_query, (self.alert_config['normal_hours'] * 60,))
                results = cursor.fetchall()

                alerts = {
                    'critical': [], 'warning': [], 'attention': [],
                    'summary': {
                        'total_waiting': len(results), 'critical_count': 0, 'warning_count': 0, 'attention_count': 0
                    }
                }

                for row in results:
                    horas = row['horas_espera'] or 0
                    alert_data = {
                        'patente': row['patente'],
                        'planilla': row.get('planilla', ''),
                        'deposito_destino': row.get('deposito_destino', ''),
                        'horas_espera': horas,
                        'estado_descarga': 'EN_PROCESO',
                        'inicio_espera': 'Calculado automáticamente',
                        'status': row.get('status', ''),
                        'alert_level': 'CRITICAL' if horas >= 48 else 'WARNING' if horas >= 8 else 'ATTENTION'
                    }

                    if horas >= self.alert_config['critical_hours']:
                        alerts['critical'].append(alert_data)
                        alerts['summary']['critical_count'] += 1
                    elif horas >= self.alert_config['warning_hours']:
                        alerts['warning'].append(alert_data)
                        alerts['summary']['warning_count'] += 1
                    else:
                        alerts['attention'].append(alert_data)
                        alerts['summary']['attention_count'] += 1

                return alerts

        except Exception as e:
            logger.error(f"Error generando alertas: {e}")
            return {
                'critical': [], 'warning': [], 'attention': [],
                'summary': {'total_waiting': 0, 'critical_count': 0, 'warning_count': 0, 'attention_count': 0}
            }

    def get_alerts_summary_complete(self):
        """Obtiene resumen completo de alertas desde BD"""
        try:
            alerts = self.generate_waiting_alerts_complete()

            if alerts and 'summary' in alerts:
                return alerts['summary']
            else:
                return {
                    'total_waiting': 0, 'critical_count': 0, 'warning_count': 0, 'attention_count': 0
                }
        except Exception as e:
            logger.error(f"Error obteniendo resumen de alertas: {e}")
            return {}

    def get_active_alerts_complete(self):
        """Obtiene alertas activas completas desde BD"""
        try:
            alerts = self.generate_waiting_alerts_complete()

            if not alerts:
                return []

            all_alerts = []
            for level, alerts_list in alerts.items():
                if level != 'summary' and isinstance(alerts_list, list):
                    for alert in alerts_list:
                        alert['alert_level'] = level.upper()
                        all_alerts.append(alert)

            all_alerts.sort(key=lambda x: x.get('horas_espera', 0), reverse=True)
            return all_alerts

        except Exception as e:
            logger.error(f"Error obteniendo alertas completas: {e}")
            return []

    def get_critical_alerts_complete(self):
        """Obtiene solo alertas críticas desde BD"""
        try:
            alerts = self.get_active_alerts_complete()
            return [alert for alert in alerts if alert.get('alert_level') == 'CRITICAL']
        except Exception as e:
            logger.error(f"Error obteniendo alertas críticas: {e}")
            return []

    def get_dashboard_stats_complete(self):
        """Obtiene estadísticas completas para dashboard desde BD y datos reales"""
        try:
            trucks_data = self.get_all_trucks_status_complete()

            if not trucks_data:
                return {
                    'total_camiones': 0, 'en_transito': 0, 'en_descarga': 0, 'alertas_criticas': 0,
                    'alertas_warning': 0, 'promedio_progreso': 0, 'geocercas': {}, 'estados_distribucion': {}
                }

            # Calcular estadísticas completas
            total_camiones = len(trucks_data)
            en_transito = len([t for t in trucks_data if t['estado_entrega'] == 'EN_TRANSITO'])
            en_descarga = len([t for t in trucks_data if 'DESCARGA' in t['estado_entrega']])
            alertas_criticas = len([t for t in trucks_data if t['alert_level'] == 'CRITICAL'])
            alertas_warning = len([t for t in trucks_data if t['alert_level'] == 'WARNING'])

            promedio_progreso = sum(t['porcentaje_entrega'] for t in trucks_data) / total_camiones

            # Estadísticas por geocerca
            geocercas_stats = {
                'en_docks': len([t for t in trucks_data if t['en_docks'] != 'NO']),
                'en_track_trace': len([t for t in trucks_data if t['en_track_trace'] != 'NO']),
                'en_cbn': len([t for t in trucks_data if t['en_cbn'] != 'NO']),
                'en_ciudades': len([t for t in trucks_data if t['en_ciudades'] != 'NO'])
            }

            # Distribución por estado
            estados_count = {}
            for truck in trucks_data:
                estado = truck['estado_entrega']
                estados_count[estado] = estados_count.get(estado, 0) + 1

            return {
                'total_camiones': total_camiones,
                'en_transito': en_transito,
                'en_descarga': en_descarga,
                'alertas_criticas': alertas_criticas,
                'alertas_warning': alertas_warning,
                'promedio_progreso': round(promedio_progreso, 2),
                'geocercas': geocercas_stats,
                'estados_distribucion': estados_count,
                'ultima_actualizacion': datetime.now().isoformat()
            }

        except Exception as e:
            logger.error(f"Error obteniendo estadísticas completas: {e}")
            return {}

    def get_geocercas_distribution(self):
        """Obtiene distribución real de camiones por geocerca"""
        try:
            trucks_data = self.get_all_trucks_status_complete()

            distribution = {
                'docks': len([t for t in trucks_data if t['en_docks'] != 'NO']),
                'track_and_trace': len([t for t in trucks_data if t['en_track_trace'] != 'NO']),
                'cbn': len([t for t in trucks_data if t['en_cbn'] != 'NO']),
                'ciudades': len([t for t in trucks_data if t['en_ciudades'] != 'NO']),
                'total': len(trucks_data)
            }

            return distribution
        except Exception as e:
            logger.error(f"Error obteniendo distribución de geocercas: {e}")
            return {}

    def get_geocercas_status(self):
        """Obtiene estado real de todas las geocercas"""
        try:
            trucks_data = self.get_all_trucks_status_complete()
            geocercas_status = []

            for grupo, geocercas_lista in self.geocercas.items():
                for geocerca in geocercas_lista:
                    camiones_dentro = 0
                    for truck in trucks_data:
                        geocerca_field = f'en_{grupo.lower().replace(" ", "_")}'
                        if truck.get(geocerca_field, 'NO') != 'NO':
                            if geocerca['nombre'].upper() in truck[geocerca_field].upper():
                                camiones_dentro += 1

                    geocercas_status.append({
                        'grupo': grupo,
                        'nombre': geocerca['nombre'],
                        'activa': geocerca['polygon'] is not None,
                        'camiones_dentro': camiones_dentro
                    })

            return geocercas_status
        except Exception as e:
            logger.error(f"Error obteniendo estado de geocercas: {e}")
            return []

    def generate_excel_report_complete(self):
        """Genera reporte Excel completo con múltiples hojas y colores"""
        try:
            # Asegurar que tenemos datos actuales
            if not self.results_data:
                self.results_data = self.get_all_trucks_status_complete()

            # Generar alertas
            alerts = self.generate_waiting_alerts_complete()

            # Usar la función de generación completa
            return self._generate_excel_report_with_alerts_complete(alerts)

        except Exception as e:
            logger.error(f"Error generando reporte Excel completo: {e}")
            return None

    def _generate_excel_report_with_alerts_complete(self, alerts: Dict):
        """Genera Excel completo con múltiples hojas y colores automáticos"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'tracking_completo_{timestamp}.xlsx'

            # Crear DataFrame con los resultados
            df = pd.DataFrame(self.results_data)

            # Crear archivo Excel con múltiples hojas
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # HOJA 1: Tracking Completo con todos los datos
                df.to_excel(writer, sheet_name='Tracking_Completo', index=False)

                # HOJA 2: Consolidado (datos únicos ordenados)
                self._create_consolidado_sheet_complete(writer, df)

                # HOJA 3: Alertas de espera (si existen)
                if alerts and alerts['summary']['total_waiting'] > 0:
                    all_alerts = alerts['critical'] + alerts['warning'] + alerts['attention']
                    df_alerts = pd.DataFrame(all_alerts)
                    df_alerts.to_excel(writer, sheet_name='🚨_ALERTAS_ESPERA', index=False)

                    # HOJA 4: Resumen de alertas
                    summary_data = [
                        {'Nivel': '🚨 CRÍTICO (>48h)', 'Cantidad': alerts['summary']['critical_count'],
                         'Descripción': 'Camiones esperando más de 48 horas'},
                        {'Nivel': '⚠️ ADVERTENCIA (>8h)', 'Cantidad': alerts['summary']['warning_count'],
                         'Descripción': 'Camiones esperando más de 8 horas'},
                        {'Nivel': '🔔 ATENCIÓN (>4h)', 'Cantidad': alerts['summary']['attention_count'],
                         'Descripción': 'Camiones esperando más de 4 horas'},
                        {'Nivel': 'TOTAL ESPERANDO', 'Cantidad': alerts['summary']['total_waiting'],
                         'Descripción': 'Total de camiones con tiempo de espera'}
                    ]
                    df_summary = pd.DataFrame(summary_data)
                    df_summary.to_excel(writer, sheet_name='📊_Resumen_Alertas', index=False)

                # HOJA 5: Resumen por geocercas
                self._create_geocercas_summary_sheet(writer)

                # HOJA 6: Estadísticas generales
                self._create_stats_sheet(writer)

                # Aplicar formato con colores automáticos
                self._apply_excel_formatting_complete(writer.book, df, alerts)

            logger.info(f"Archivo Excel completo generado: {filename}")
            return filename

        except Exception as e:
            logger.error(f"Error generando Excel completo: {e}")
            return None

    def _create_consolidado_sheet_complete(self, writer, df):
        """Crea hoja consolidado ordenada por prioridad (como original)"""
        try:
            # Crear copia para evitar warnings
            df_consolidado = df.drop_duplicates(subset=['patente', 'planilla'], keep='last').copy()

            # Ordenamiento por prioridad (como tu original)
            estado_orden = {
                'DESCARGANDO_CONFIRMADO': 1, 'DESCARGANDO': 2, 'EN_ZONA_DESCARGA': 3,
                'EN_CENTRO_DISTRIBUCION': 4, 'EN_CIUDAD': 5, 'EN_TRANSITO': 6
            }
            alert_orden = {'CRITICAL': 1, 'WARNING': 2, 'ATTENTION': 3, 'NORMAL': 4}

            # Mapear y ordenar
            df_consolidado.loc[:, 'estado_orden'] = df_consolidado['estado_entrega'].map(estado_orden).fillna(999)
            df_consolidado.loc[:, 'alert_orden'] = df_consolidado['alert_level'].map(alert_orden).fillna(999)

            # Ordenar por: 1) Alert_level, 2) Estado_entrega, 3) Tiempo_espera (descendente)
            df_consolidado = df_consolidado.sort_values([
                'alert_orden', 'estado_orden', 'tiempo_espera_minutos'
            ], ascending=[True, True, False])

            # Remover columnas auxiliares
            df_consolidado = df_consolidado.drop(['estado_orden', 'alert_orden'], axis=1)

            # Guardar en Excel
            df_consolidado.to_excel(writer, sheet_name='Consolidado', index=False)
            logger.info(f"📊 Hoja Consolidado creada: {len(df_consolidado)} registros ordenados por prioridad")

        except Exception as e:
            logger.error(f"Error creando hoja consolidado: {e}")

    def _apply_excel_formatting_complete(self, workbook, df, alerts):
        """Aplica formato completo con colores automáticos por alert_level (como original)"""
        try:
            from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

            # Definir estilos con colores por alert_level
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)

            # Colores por nivel de alerta (como tu original)
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # NORMAL
            blue_fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')  # ATTENTION
            yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # WARNING
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # CRITICAL

            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            # Aplicar formato a hojas principales
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]

                # Formato para headers
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Aplicar colores por alert_level solo en hojas principales
                if sheet_name in ['Tracking_Completo', 'Consolidado', '🚨_ALERTAS_ESPERA']:
                    # Buscar columna alert_level
                    alert_col = None
                    for idx, cell in enumerate(worksheet[1]):
                        if cell.value and 'alert' in str(cell.value).lower():
                            alert_col = idx + 1
                            break

                    if alert_col:
                        # Aplicar colores fila por fila según alert_level
                        for row_idx in range(2, worksheet.max_row + 1):
                            alert_cell = worksheet.cell(row=row_idx, column=alert_col)
                            alert_level = str(alert_cell.value).upper() if alert_cell.value else 'NORMAL'

                            # Seleccionar color según nivel de alerta
                            if 'CRITICAL' in alert_level:
                                fill_color = red_fill
                            elif 'WARNING' in alert_level:
                                fill_color = yellow_fill
                            elif 'ATTENTION' in alert_level:
                                fill_color = blue_fill
                            else:
                                fill_color = green_fill

                            # Aplicar color a toda la fila
                            for col in range(1, worksheet.max_column + 1):
                                cell = worksheet.cell(row=row_idx, column=col)
                                cell.fill = fill_color
                                cell.border = thin_border
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        # Si no hay columna alert_level, solo bordes
                        for row_idx in range(2, worksheet.max_row + 1):
                            for col in range(1, worksheet.max_column + 1):
                                cell = worksheet.cell(row=row_idx, column=col)
                                cell.border = thin_border
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    # Para hojas de resumen, solo bordes y alineación
                    for row_idx in range(2, worksheet.max_row + 1):
                        for col in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row_idx, column=col)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center', vertical='center')

                # Ajustar ancho de columnas automáticamente
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                    # Ajustar ancho con límites
                    adjusted_width = min(max(max_length + 2, 10), 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            logger.info("🎨 Formato Excel con colores aplicado correctamente")

        except Exception as e:
            logger.error(f"Error aplicando formato con colores: {e}")

    def _create_geocercas_summary_sheet(self, writer):
        """Crea hoja con resumen de distribución por geocercas"""
        try:
            distribution = self.get_geocercas_distribution()

            geocercas_data = [
                {'Geocerca': 'DOCKS', 'Camiones': distribution.get('docks', 0),
                 'Porcentaje': f"{(distribution.get('docks', 0) / max(distribution.get('total', 1), 1)) * 100:.1f}%"},
                {'Geocerca': 'TRACK AND TRACE', 'Camiones': distribution.get('track_and_trace', 0),
                 'Porcentaje': f"{(distribution.get('track_and_trace', 0) / max(distribution.get('total', 1), 1)) * 100:.1f}%"},
                {'Geocerca': 'CBN', 'Camiones': distribution.get('cbn', 0),
                 'Porcentaje': f"{(distribution.get('cbn', 0) / max(distribution.get('total', 1), 1)) * 100:.1f}%"},
                {'Geocerca': 'CIUDADES', 'Camiones': distribution.get('ciudades', 0),
                 'Porcentaje': f"{(distribution.get('ciudades', 0) / max(distribution.get('total', 1), 1)) * 100:.1f}%"},
                {'Geocerca': 'TOTAL', 'Camiones': distribution.get('total', 0), 'Porcentaje': '100.0%'}
            ]

            df_geocercas = pd.DataFrame(geocercas_data)
            df_geocercas.to_excel(writer, sheet_name='Resumen_Geocercas', index=False)

        except Exception as e:
            logger.error(f"Error creando hoja resumen geocercas: {e}")

    def _create_stats_sheet(self, writer):
        """Crea hoja con estadísticas generales del sistema"""
        try:
            stats = self.get_dashboard_stats_complete()

            stats_data = [
                {'Métrica': 'Total Camiones', 'Valor': stats.get('total_camiones', 0)},
                {'Métrica': 'En Tránsito', 'Valor': stats.get('en_transito', 0)},
                {'Métrica': 'En Descarga', 'Valor': stats.get('en_descarga', 0)},
                {'Métrica': 'Alertas Críticas', 'Valor': stats.get('alertas_criticas', 0)},
                {'Métrica': 'Alertas Warning', 'Valor': stats.get('alertas_warning', 0)},
                {'Métrica': 'Promedio Progreso (%)', 'Valor': stats.get('promedio_progreso', 0)},
                {'Métrica': 'Generado', 'Valor': datetime.now().strftime('%Y-%m-%d %H:%M:%S')},
                {'Métrica': 'Sistema', 'Valor': 'Tracking CBN v2.0'}
            ]

            df_stats = pd.DataFrame(stats_data)
            df_stats.to_excel(writer, sheet_name='Estadísticas', index=False)

        except Exception as e:
            logger.error(f"Error creando hoja estadísticas: {e}")

    def _create_stats_sheet(self, writer):
        """Crea hoja con estadísticas generales del sistema"""
        try:
            stats = self.get_dashboard_stats_complete()

            stats_data = [
                {'Métrica': 'Total Camiones', 'Valor': stats.get('total_camiones', 0)},
                {'Métrica': 'En Tránsito', 'Valor': stats.get('en_transito', 0)},
                {'Métrica': 'En Descarga', 'Valor': stats.get('en_descarga', 0)},
                {'Métrica': 'Alertas Críticas', 'Valor': stats.get('alertas_criticas', 0)},
                {'Métrica': 'Alertas Warning', 'Valor': stats.get('alertas_warning', 0)},
                {'Métrica': 'Promedio Progreso (%)', 'Valor': stats.get('promedio_progreso', 0)},
                {'Métrica': 'Generado', 'Valor': datetime.now().strftime('%Y-%m-%d %H:%M:%S')},
                {'Métrica': 'Sistema', 'Valor': 'Tracking CBN v2.0'}
            ]

            df_stats = pd.DataFrame(stats_data)
            df_stats.to_excel(writer, sheet_name='Estadísticas', index=False)

        except Exception as e:
            logger.error(f"Error creando hoja estadísticas: {e}")

    def process_all_trucks_complete(self):
        """Procesa todos los camiones con funcionalidad completa (igual que original)"""
        start_time = time.time()
        logger.info("🚀 Iniciando procesamiento completo con geocercas y alertas...")

        try:
            with self.processing_lock:
                # Limpiar resultados anteriores
                self.results_data = []

                # Cargar geocercas si no están cargadas
                if not self.geocercas:
                    self.load_geocercas()

                # Actualizar datos históricos para camiones existentes
                if self.historical_data:
                    self.update_historical_waiting_times()

                # Obtener y procesar camiones
                trucks = self.get_trucks_in_transit()
                if not trucks:
                    logger.info("No hay camiones en tránsito")
                    return

                # Obtener todas las ubicaciones
                all_locations = self.get_all_trucks_locations_parallel(trucks)

                # Procesar cada camión
                processed = 0
                errors = 0

                for truck in trucks:
                    try:
                        patente = truck['patente']
                        planilla = truck['planilla']
                        deposito_destino = truck.get('deposito_destino', '')

                        location = all_locations.get(patente)

                        if location and location['latitude'] and location['longitude']:
                            # Verificar geocercas
                            geocerca_status = self.check_point_in_geocercas(
                                location['latitude'], location['longitude'], deposito_destino
                            )

                            # Calcular progreso
                            porcentaje_entrega, estado_entrega = self.calculate_delivery_progress(
                                geocerca_status, deposito_destino
                            )

                            # Calcular tiempo de espera
                            tiempo_espera_minutos, inicio_espera_str, estado_descarga, alert_level = \
                                self.calculate_waiting_time_for_discharge(truck, geocerca_status, estado_entrega)

                            # Guardar en BD
                            self._save_truck_tracking_complete(
                                truck, location, geocerca_status, porcentaje_entrega, estado_entrega,
                                tiempo_espera_minutos, estado_descarga, inicio_espera_str
                            )

                            # Preparar datos para Excel
                            excel_row = {
                                'patente': patente,
                                'planilla': planilla,
                                'status': truck.get('status', ''),
                                'deposito_origen': truck.get('deposito_origen', ''),
                                'deposito_destino': deposito_destino,
                                'producto': truck.get('producto', ''),
                                'cod_producto': truck.get('cod_producto', ''),
                                'salida': truck.get('salida', ''),
                                'fecha_salida': truck.get('fecha_salida', ''),
                                'hora_salida': self._adjust_time_utc_minus_4(truck.get('hora_salida', '')),
                                'fecha_llegada': truck.get('fecha_llegada', ''),
                                'hora_llegada': self._adjust_time_utc_minus_4(truck.get('hora_llegada', '')),
                                'latitude': location.get('latitude'),
                                'longitude': location.get('longitude'),
                                'velocidad_kmh': location.get('speed', 0),
                                'timestamp': location.get('timestamp', ''),
                                'en_docks': geocerca_status['DOCKS'],
                                'en_track_trace': geocerca_status['TRACK AND TRACE'],
                                'en_cbn': geocerca_status['CBN'],
                                'en_ciudades': geocerca_status['CIUDADES'],
                                'porcentaje_entrega': porcentaje_entrega,
                                'estado_entrega': estado_entrega,
                                'tiempo_espera_minutos': tiempo_espera_minutos,
                                'tiempo_espera_horas': round(tiempo_espera_minutos / 60,
                                                             2) if tiempo_espera_minutos > 0 else 0,
                                'estado_descarga': estado_descarga,
                                'alert_level': alert_level,
                                'inicio_espera': inicio_espera_str,
                                'fecha_proceso': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            }
                            self.results_data.append(excel_row)

                            # Log de progreso
                            tiempo_espera_str = ""
                            if tiempo_espera_minutos > 0:
                                horas = tiempo_espera_minutos // 60
                                minutos = tiempo_espera_minutos % 60
                                alert_emoji = {"CRITICAL": "🚨", "WARNING": "⚠️", "ATTENTION": "🔔"}.get(alert_level, "⏰")
                                tiempo_espera_str = f" {alert_emoji} Esperando: {horas}h {minutos}m"

                            # Log de resultados
                            in_geocerca = [f"{geo}: {status}" for geo, status in geocerca_status.items() if
                                           status != 'NO']
                            geocerca_str = ', '.join(in_geocerca) if in_geocerca else 'En tránsito libre'

                            logger.info(
                                f"✅ {patente}: {porcentaje_entrega}% ({estado_entrega}) - {geocerca_str}{tiempo_espera_str}")
                            processed += 1
                        else:
                            logger.warning(f"⚠️ {patente}: Sin ubicación válida")
                            errors += 1

                    except Exception as e:
                        logger.error(f"❌ Error procesando {truck.get('patente', 'UNKNOWN')}: {e}")
                        errors += 1

                elapsed_time = time.time() - start_time
                logger.info(
                    f"🏁 Procesamiento completo terminado en {elapsed_time:.2f}s: {processed} exitosos, {errors} errores")

                # Generar alertas finales
                alerts = self.generate_waiting_alerts_complete()
                if alerts['summary']['total_waiting'] > 0:
                    logger.info(f"📊 ALERTAS GENERADAS:")
                    logger.info(f"   🚨 Críticas (>48h): {alerts['summary']['critical_count']}")
                    logger.info(f"   ⚠️ Advertencias (>8h): {alerts['summary']['warning_count']}")
                    logger.info(f"   🔔 Atención (>4h): {alerts['summary']['attention_count']}")

                # Actualizar timestamp de procesamiento
                self.last_processing_time = datetime.now()

                # Limpiar cache para forzar actualización
                self.cache['last_update'] = None

                logger.info("✅ Procesamiento completo finalizado exitosamente")

        except Exception as e:
            logger.error(f"Error en procesamiento completo: {e}")
            raise

    def update_historical_waiting_times(self):
        """Actualiza tiempos de espera usando datos históricos del Excel"""
        if not self.historical_data:
            return

        try:
            with self.target_connection.cursor() as cursor:
                updated_count = 0

                for patente, hist_data in self.historical_data.items():
                    # Buscar registros que necesiten actualización
                    check_query = """
                    SELECT id, patente, planilla FROM truck_tracking 
                    WHERE patente = %s AND (inicio_espera_descarga IS NULL OR inicio_espera_descarga = '')
                    AND status = 'SALIDA'
                    """

                    cursor.execute(check_query, (patente,))
                    trucks_to_update = cursor.fetchall()

                    for truck in trucks_to_update:
                        inicio_espera = hist_data['primera_entrada_descarga']
                        tiempo_espera = datetime.now() - inicio_espera
                        tiempo_espera_minutos = int(tiempo_espera.total_seconds() / 60)

                        # Determinar alert_level basado en tiempo histórico
                        alert_level = self._get_alert_level(tiempo_espera_minutos)

                        update_query = """
                        UPDATE truck_tracking SET
                            inicio_espera_descarga = %s,
                            tiempo_espera_minutos = %s,
                            estado_descarga = 'HISTORICO_EXCEL',
                            alert_level = %s
                        WHERE id = %s
                        """

                        cursor.execute(update_query, (
                            inicio_espera.strftime('%Y-%m-%d %H:%M:%S'),
                            tiempo_espera_minutos,
                            alert_level,
                            truck['id']
                        ))
                        updated_count += 1

                        logger.info(
                            f"📅 Histórico actualizado para {patente}: {tiempo_espera_minutos // 60}h {tiempo_espera_minutos % 60}m esperando")

                self.target_connection.commit()
                logger.info(f"✅ Actualizados {updated_count} registros con datos históricos")

        except Exception as e:
            logger.error(f"Error actualizando tiempos históricos: {e}")
            self.target_connection.rollback()

    def disconnect_databases(self):
        """Desconecta de ambas bases de datos"""
        try:
            if self.source_connection:
                self.source_connection.close()
                logger.info("✅ Conexión BD origen cerrada")
        except Exception as e:
            logger.warning(f"⚠️ Error cerrando conexión origen: {e}")

        try:
            if self.target_connection:
                self.target_connection.close()
                logger.info("✅ Conexión BD destino cerrada")
        except Exception as e:
            logger.warning(f"⚠️ Error cerrando conexión destino: {e}")

    def get_system_health(self):
        """Obtiene estado de salud del sistema"""
        try:
            health = {
                'timestamp': datetime.now().isoformat(),
                'databases': {
                    'source': self.source_connection is not None and self.source_connection.open,
                    'target': self.target_connection is not None and self.target_connection.open
                },
                'geocercas': {
                    'loaded': len(self.geocercas) > 0,
                    'count': len(self.geocercas),
                    'groups': list(self.geocercas.keys()) if self.geocercas else []
                },
                'historical_data': {
                    'loaded': len(self.historical_data) > 0,
                    'trucks_count': len(self.historical_data)
                },
                'cache': {
                    'trucks_cached': len(self.cache['trucks_data']),
                    'last_update': self.cache['last_update'].isoformat() if self.cache['last_update'] else None
                },
                'last_processing': self.last_processing_time.isoformat() if self.last_processing_time else None,
                'is_processing': self.processing_lock.locked()
            }
            return health
        except Exception as e:
            logger.error(f"Error obteniendo salud del sistema: {e}")
            return {'error': str(e)}

    def clear_cache(self):
        """Limpia el cache del sistema"""
        try:
            self.cache = {
                'trucks_data': [],
                'alerts': {},
                'stats': {},
                'last_update': None
            }
            logger.info("🧹 Cache limpiado correctamente")
            return True
        except Exception as e:
            logger.error(f"Error limpiando cache: {e}")
            return False

    def force_reload_geocercas(self):
        """Fuerza recarga de geocercas desde Excel"""
        try:
            self.geocercas = {}
            success = self.load_geocercas()
            if success:
                logger.info("🔄 Geocercas recargadas exitosamente")
            return success
        except Exception as e:
            logger.error(f"Error recargando geocercas: {e}")
            return False

    def force_reload_historical_data(self):
        """Fuerza recarga de datos históricos desde Excel"""
        try:
            self.historical_data = {}
            success = self.load_historical_data()
            if success:
                logger.info("🔄 Datos históricos recargados exitosamente")
            return success
        except Exception as e:
            logger.error(f"Error recargando datos históricos: {e}")
            return False

    def get_processing_stats(self):
        """Obtiene estadísticas de procesamiento"""
        try:
            with self.target_connection.cursor() as cursor:
                # Estadísticas básicas
                stats_query = """
                SELECT 
                    COUNT(*) as total_registros,
                    COUNT(DISTINCT patente) as total_camiones,
                    AVG(porcentaje_entrega) as promedio_progreso,
                    AVG(tiempo_espera_minutos) as promedio_espera_minutos,
                    MIN(ultima_actualizacion) as primer_registro,
                    MAX(ultima_actualizacion) as ultimo_registro
                FROM truck_tracking
                WHERE status = 'SALIDA'
                """
                cursor.execute(stats_query)
                stats = cursor.fetchone()

                # Estadísticas por alert_level
                alerts_query = """
                SELECT alert_level, COUNT(*) as cantidad
                FROM truck_tracking
                WHERE status = 'SALIDA'
                GROUP BY alert_level
                """
                cursor.execute(alerts_query)
                alerts_stats = {row['alert_level']: row['cantidad'] for row in cursor.fetchall()}

                # Estadísticas por estado_entrega
                estados_query = """
                SELECT estado_entrega, COUNT(*) as cantidad
                FROM truck_tracking
                WHERE status = 'SALIDA'
                GROUP BY estado_entrega
                """
                cursor.execute(estados_query)
                estados_stats = {row['estado_entrega']: row['cantidad'] for row in cursor.fetchall()}

                return {
                    'general': stats,
                    'por_alert_level': alerts_stats,
                    'por_estado_entrega': estados_stats,
                    'timestamp': datetime.now().isoformat()
                }

        except Exception as e:
            logger.error(f"Error obteniendo estadísticas de procesamiento: {e}")
            return {}

    def backup_current_data(self):
        """Crea backup de los datos actuales en CSV"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_filename = f'backup_tracking_{timestamp}.csv'

            with self.target_connection.cursor() as cursor:
                backup_query = """
                SELECT * FROM truck_tracking 
                WHERE status = 'SALIDA'
                ORDER BY ultima_actualizacion DESC
                """
                cursor.execute(backup_query)
                results = cursor.fetchall()

                if results:
                    df_backup = pd.DataFrame(results)
                    df_backup.to_csv(backup_filename, index=False, encoding='utf-8')
                    logger.info(f"💾 Backup creado: {backup_filename} ({len(results)} registros)")
                    return backup_filename
                else:
                    logger.warning("⚠️ No hay datos para hacer backup")
                    return None

        except Exception as e:
            logger.error(f"Error creando backup: {e}")
            return None

    def cleanup_old_records(self, days_old=30):
        """Limpia registros antiguos de la BD (mayor a X días)"""
        try:
            with self.target_connection.cursor() as cursor:
                cleanup_query = """
                DELETE FROM truck_tracking 
                WHERE ultima_actualizacion < DATE_SUB(NOW(), INTERVAL %s DAY)
                AND status != 'SALIDA'
                """
                cursor.execute(cleanup_query, (days_old,))
                deleted_count = cursor.rowcount
                self.target_connection.commit()

                logger.info(f"🧹 Limpieza completada: {deleted_count} registros antiguos eliminados")
                return deleted_count

        except Exception as e:
            logger.error(f"Error en limpieza de registros: {e}")
            self.target_connection.rollback()
            return 0

    def test_api_connection(self):
        """Prueba la conexión con la API de Boltrack"""
        try:
            headers = {
                'token': self.config['api']['token'],
                'Content-Type': 'application/json'
            }

            # Hacer una petición simple para probar
            response = requests.get(
                f"{self.config['api']['base_url']}/ultimaubicaciontodos",
                headers=headers,
                timeout=10
            )

            api_status = {
                'status_code': response.status_code,
                'response_time_ms': response.elapsed.total_seconds() * 1000,
                'success': response.status_code == 200,
                'timestamp': datetime.now().isoformat()
            }

            if response.status_code == 200:
                data = response.json()
                api_status['vehicles_count'] = len(data) if isinstance(data, list) else 0
                logger.info(
                    f"✅ API funcionando: {api_status['vehicles_count']} vehículos, {api_status['response_time_ms']:.1f}ms")
            else:
                logger.warning(f"⚠️ API respondió con código {response.status_code}")

            return api_status

        except Exception as e:
            logger.error(f"❌ Error probando API: {e}")
            return {
                'success': False,
                'error': str(e),
                'timestamp': datetime.now().isoformat()
            }

    def get_cache_info(self):
        """Obtiene información detallada del cache"""
        return {
            'trucks_data': {
                'count': len(self.cache['trucks_data']),
                'last_update': self.cache['last_update'].isoformat() if self.cache['last_update'] else None,
                'size_bytes': len(str(self.cache['trucks_data']))
            },
            'geocercas': {
                'groups_loaded': len(self.geocercas),
                'total_geocercas': sum(len(geocercas_list) for geocercas_list in self.geocercas.values()),
                'groups': list(self.geocercas.keys())
            },
            'historical_data': {
                'trucks_count': len(self.historical_data),
                'size_bytes': len(str(self.historical_data))
            },
            'results_data': {
                'count': len(self.results_data),
                'size_bytes': len(str(self.results_data))
            }
        }

    def _update_table_structure(self):
        """Actualiza automáticamente la estructura de la tabla (compatible MySQL)"""
        try:
            with self.target_connection.cursor() as cursor:
                # Primero verificar qué columnas existen
                cursor.execute("SHOW COLUMNS FROM truck_tracking")
                existing_columns = [row['Field'] for row in cursor.fetchall()]

                # Columnas que necesitamos agregar
                columns_needed = {
                    'geocerca_docks': "VARCHAR(255) DEFAULT 'NO'",
                    'geocerca_track_trace': "VARCHAR(255) DEFAULT 'NO'",
                    'geocerca_cbn': "VARCHAR(255) DEFAULT 'NO'",
                    'geocerca_ciudades': "VARCHAR(255) DEFAULT 'NO'",
                    'inicio_espera_descarga': "DATETIME NULL",
                    'alert_level': "VARCHAR(20) DEFAULT 'NORMAL'",
                    'estado_descarga': "VARCHAR(50) DEFAULT 'NO_INICIADO'",
                    'primera_deteccion': "DATETIME DEFAULT CURRENT_TIMESTAMP",
                    'cod': "VARCHAR(255) NULL",
                    'deposito_origen': "VARCHAR(255) NULL",
                    'cod_destino': "VARCHAR(255) NULL",
                    'fecha_salida': "DATE NULL",
                    'hora_salida': "TIME NULL",
                    'fecha_llegada': "DATE NULL",
                    'hora_llegada': "TIME NULL",
                    'cod_producto': "VARCHAR(255) NULL",
                    'producto': "VARCHAR(255) NULL",
                    'direccion': "INT NULL",
                    'salida': "INT NULL"
                }

                # Agregar solo las columnas que faltan
                added_count = 0
                for column_name, column_definition in columns_needed.items():
                    if column_name not in existing_columns:
                        try:
                            alter_query = f"ALTER TABLE truck_tracking ADD COLUMN {column_name} {column_definition}"
                            cursor.execute(alter_query)
                            logger.info(f"✅ Columna agregada: {column_name}")
                            added_count += 1
                        except Exception as e:
                            logger.error(f"❌ Error agregando {column_name}: {e}")

                self.target_connection.commit()
                logger.info(f"✅ Estructura actualizada: {added_count} columnas agregadas")

        except Exception as e:
            logger.error(f"Error actualizando tabla: {e}")
            self.target_connection.rollback()

    # Agregar a truck_tracking_web_complete.py
    def get_geocercas_for_map(self):
        """Convierte geocercas a formato GeoJSON para visualización"""
        geocercas_geojson = {
            "type": "FeatureCollection",
            "features": []
        }

        for grupo, geocercas_lista in self.geocercas.items():
            for geocerca in geocercas_lista:
                if geocerca['polygon']:
                    # Convertir Polygon a GeoJSON
                    coords = list(geocerca['polygon'].exterior.coords)
                    feature = {
                        "type": "Feature",
                        "geometry": {
                            "type": "Polygon",
                            "coordinates": [coords]
                        },
                        "properties": {
                            "grupo": grupo,
                            "nombre": geocerca['nombre'],
                            "color": self._get_geocerca_color(grupo)
                        }
                    }
                    geocercas_geojson["features"].append(feature)

        return geocercas_geojson

    def _get_geocerca_color(self, grupo):
        """Asigna colores a grupos de geocercas"""
        colors = {
            'DOCKS': '#ff6b6b',
            'TRACK AND TRACE': '#4ecdc4',
            'CBN': '#45b7d1',
            'CIUDADES': '#96ceb4'
        }
        return colors.get(grupo, '#95a5a6')

    # ===============================
    # SISTEMA DE ALERTAS AVANZADO
    # ===============================

    def get_alerts_dashboard_data(self):
        """Obtiene datos completos para el dashboard de alertas"""
        try:
            trucks = self.get_all_trucks_status_complete()
            alerts_data = self.generate_waiting_alerts_complete()

            # Estadísticas generales
            total_trucks = len(trucks)
            total_alerts = alerts_data['summary']['total_waiting']

            # Alertas por depósito destino
            alerts_by_destination = {}
            for truck in trucks:
                if truck.get('alert_level') != 'NORMAL':
                    dest = truck.get('deposito_destino', 'Sin destino')
                    if dest not in alerts_by_destination:
                        alerts_by_destination[dest] = {
                            'critical': 0, 'warning': 0, 'attention': 0, 'total': 0
                        }

                    level = truck.get('alert_level', 'NORMAL').lower()
                    if level in alerts_by_destination[dest]:
                        alerts_by_destination[dest][level] += 1
                    alerts_by_destination[dest]['total'] += 1

            # Tendencias por hora (últimas 24 horas simuladas)
            hourly_trends = self._generate_hourly_alert_trends()

            # Alertas críticas con detalles completos
            critical_alerts_detailed = []
            for truck in trucks:
                if truck.get('alert_level') == 'CRITICAL':
                    alert_detail = {
                        'patente': truck['patente'],
                        'planilla': truck.get('planilla', ''),
                        'deposito_destino': truck.get('deposito_destino', ''),
                        'tiempo_espera_horas': truck.get('tiempo_espera_horas', 0),
                        'estado_entrega': truck.get('estado_entrega', ''),
                        'ubicacion': {
                            'latitude': truck.get('latitude'),
                            'longitude': truck.get('longitude')
                        },
                        'geocercas_activas': self._get_active_geocercas(truck),
                        'inicio_espera': truck.get('inicio_espera', ''),
                        'producto': truck.get('producto', ''),
                        'velocidad_kmh': truck.get('velocidad_kmh', 0),
                        'escalamiento_requerido': truck.get('tiempo_espera_horas', 0) > 72,
                        'prioridad': self._calculate_alert_priority(truck)
                    }
                    critical_alerts_detailed.append(alert_detail)

            # Ordenar por prioridad
            critical_alerts_detailed.sort(key=lambda x: x['prioridad'], reverse=True)

            # Resumen ejecutivo
            executive_summary = {
                'total_camiones_problema': total_alerts,
                'porcentaje_con_alertas': round((total_alerts / max(total_trucks, 1)) * 100, 1),
                'tiempo_espera_promedio': self._calculate_average_waiting_time(trucks),
                'deposito_mas_problematico': max(alerts_by_destination.items(),
                                                 key=lambda x: x[1]['total'])[0] if alerts_by_destination else 'N/A',
                'alertas_nuevas_ultima_hora': self._count_new_alerts_last_hour(),
                'trend_direction': self._get_alerts_trend_direction()
            }

            return {
                'summary': alerts_data['summary'],
                'alerts_by_destination': alerts_by_destination,
                'hourly_trends': hourly_trends,
                'critical_alerts_detailed': critical_alerts_detailed,
                'executive_summary': executive_summary,
                'all_alerts': alerts_data['critical'] + alerts_data['warning'] + alerts_data['attention'],
                'timestamp': datetime.now().isoformat(),
                'next_escalation': self._get_next_escalation_time(),
                'recommendations': self._generate_alert_recommendations(trucks)
            }

        except Exception as e:
            logger.error(f"Error obteniendo datos de alertas: {e}")
            return {}

    def _get_active_geocercas(self, truck):
        """Obtiene lista de geocercas activas para un camión"""
        active_geocercas = []
        geocerca_mapping = {
            'en_docks': 'DOCKS',
            'en_track_trace': 'TRACK_AND_TRACE',
            'en_cbn': 'CBN',
            'en_ciudades': 'CIUDADES'
        }

        for field, name in geocerca_mapping.items():
            if truck.get(field, 'NO') != 'NO':
                active_geocercas.append({
                    'name': name,
                    'details': truck.get(field)
                })

        return active_geocercas

    def _calculate_alert_priority(self, truck):
        """Calcula prioridad de alerta basada en múltiples factores"""
        priority = 0

        # Factor tiempo de espera (peso: 40%)
        tiempo_espera = truck.get('tiempo_espera_horas', 0)
        if tiempo_espera > 72:
            priority += 40
        elif tiempo_espera > 48:
            priority += 30
        elif tiempo_espera > 24:
            priority += 20
        else:
            priority += 10

        # Factor geocerca (peso: 30%)
        if truck.get('en_docks', 'NO') != 'NO':
            priority += 30  # En DOCKS es máxima prioridad
        elif truck.get('en_track_trace', 'NO') != 'NO':
            priority += 25
        elif truck.get('en_cbn', 'NO') != 'NO':
            priority += 15

        # Factor velocidad (peso: 20%)
        velocidad = truck.get('velocidad_kmh', 0)
        if velocidad == 0:
            priority += 20  # Parado completamente
        elif velocidad < 5:
            priority += 15

        # Factor producto (peso: 10%)
        producto = truck.get('producto', '').upper()
        if 'PREMIUM' in producto or 'ESPECIAL' in producto:
            priority += 10
        elif 'URGENTE' in producto:
            priority += 8

        return min(priority, 100)  # Máximo 100

    def _generate_hourly_alert_trends(self):
        """Genera tendencias por hora (simuladas por ahora)"""
        import random
        trends = []
        base_time = datetime.now().replace(minute=0, second=0, microsecond=0)

        for i in range(24):
            hour_time = base_time - timedelta(hours=i)
            # Simular datos realistas basados en horarios
            hour = hour_time.hour

            # Más alertas durante horarios de trabajo
            if 8 <= hour <= 18:
                base_alerts = random.randint(3, 8)
            elif 6 <= hour <= 22:
                base_alerts = random.randint(1, 5)
            else:
                base_alerts = random.randint(0, 3)

            trends.append({
                'hour': hour_time.strftime('%H:%M'),
                'timestamp': hour_time.isoformat(),
                'critical': max(0, base_alerts - 4),
                'warning': min(base_alerts, 3),
                'attention': min(base_alerts, 2),
                'total': base_alerts
            })

        return list(reversed(trends))  # Orden cronológico

    def _calculate_average_waiting_time(self, trucks):
        """Calcula tiempo promedio de espera"""
        waiting_trucks = [t for t in trucks if t.get('tiempo_espera_horas', 0) > 0]
        if not waiting_trucks:
            return 0

        total_waiting = sum(t.get('tiempo_espera_horas', 0) for t in waiting_trucks)
        return round(total_waiting / len(waiting_trucks), 1)

    def _count_new_alerts_last_hour(self):
        """Cuenta alertas nuevas en la última hora (simulado)"""
        # En producción, esto consultaría el histórico de la BD
        import random
        return random.randint(0, 5)

    def _get_alerts_trend_direction(self):
        """Determina dirección de tendencia de alertas"""
        # En producción, compararía con períodos anteriores
        import random
        directions = ['increasing', 'decreasing', 'stable']
        return random.choice(directions)

    def _get_next_escalation_time(self):
        """Calcula próximo tiempo de escalación"""
        return (datetime.now() + timedelta(hours=2)).strftime('%H:%M')

    def _generate_alert_recommendations(self, trucks):
        """Genera recomendaciones automáticas"""
        recommendations = []

        critical_count = len([t for t in trucks if t.get('alert_level') == 'CRITICAL'])
        warning_count = len([t for t in trucks if t.get('alert_level') == 'WARNING'])

        if critical_count > 3:
            recommendations.append({
                'type': 'urgent',
                'title': 'Múltiples Alertas Críticas',
                'description': f'{critical_count} camiones con más de 48h de espera',
                'action': 'Activar protocolo de escalación inmediata',
                'priority': 'high'
            })

        if warning_count > 5:
            recommendations.append({
                'type': 'attention',
                'title': 'Tendencia Creciente de Esperas',
                'description': f'{warning_count} camiones esperando más de 8 horas',
                'action': 'Revisar capacidad de descarga en centros',
                'priority': 'medium'
            })

        # Análisis por geocerca
        docks_trucks = len([t for t in trucks if t.get('en_docks', 'NO') != 'NO'])
        if docks_trucks > 2:
            recommendations.append({
                'type': 'operational',
                'title': 'Congestión en DOCKS',
                'description': f'{docks_trucks} camiones esperando en muelles',
                'action': 'Optimizar asignación de muelles',
                'priority': 'medium'
            })

        return recommendations

    def get_alert_configurations(self):
        """Obtiene configuraciones de alertas"""
        return {
            'thresholds': {
                'normal_hours': self.alert_config['normal_hours'],
                'warning_hours': self.alert_config['warning_hours'],
                'critical_hours': self.alert_config['critical_hours']
            },
            'notification_settings': {
                'email_enabled': True,
                'sms_enabled': False,
                'whatsapp_enabled': False,
                'auto_escalation': True,
                'escalation_hours': 72
            },
            'monitoring': {
                'refresh_interval': 300,  # 5 minutos
                'dashboard_refresh': 30,  # 30 segundos
                'critical_notification_delay': 60  # 1 minuto
            }
        }

    def update_alert_configurations(self, new_config):
        """Actualiza configuraciones de alertas"""
        try:
            if 'thresholds' in new_config:
                thresholds = new_config['thresholds']
                if 'normal_hours' in thresholds:
                    self.alert_config['normal_hours'] = float(thresholds['normal_hours'])
                if 'warning_hours' in thresholds:
                    self.alert_config['warning_hours'] = float(thresholds['warning_hours'])
                if 'critical_hours' in thresholds:
                    self.alert_config['critical_hours'] = float(thresholds['critical_hours'])

            logger.info(f"Configuración de alertas actualizada: {self.alert_config}")
            return True

        except Exception as e:
            logger.error(f"Error actualizando configuración de alertas: {e}")
            return False

    def generate_alert_report(self, start_date=None, end_date=None):
        """Genera reporte de alertas para un período"""
        try:
            if not start_date:
                start_date = datetime.now() - timedelta(days=7)
            if not end_date:
                end_date = datetime.now()

            # En producción, esto consultaría la BD histórica
            # Por ahora, generamos datos de ejemplo

            with self.target_connection.cursor() as cursor:
                report_query = """
                SELECT 
                    DATE(ultima_actualizacion) as fecha,
                    alert_level,
                    COUNT(*) as cantidad,
                    AVG(tiempo_espera_minutos) as tiempo_promedio,
                    MAX(tiempo_espera_minutos) as tiempo_maximo
                FROM truck_tracking 
                WHERE ultima_actualizacion BETWEEN %s AND %s
                AND alert_level != 'NORMAL'
                GROUP BY DATE(ultima_actualizacion), alert_level
                ORDER BY fecha DESC, alert_level
                """

                cursor.execute(report_query, (start_date, end_date))
                results = cursor.fetchall()

                # Procesar resultados para el reporte
                report_data = {
                    'period': {
                        'start': start_date.strftime('%Y-%m-%d'),
                        'end': end_date.strftime('%Y-%m-%d')
                    },
                    'summary': {
                        'total_alerts': len(results),
                        'avg_daily_alerts': len(results) / max((end_date - start_date).days, 1)
                    },
                    'daily_breakdown': results,
                    'trends': self._analyze_alert_trends(results),
                    'generated_at': datetime.now().isoformat()
                }

                return report_data

        except Exception as e:
            logger.error(f"Error generando reporte de alertas: {e}")
            return {}

    def _analyze_alert_trends(self, data):
        """Analiza tendencias en los datos de alertas"""
        if not data:
            return {'trend': 'no_data', 'direction': 'stable'}

        # Análisis simple de tendencia
        recent_alerts = sum(row['cantidad'] for row in data[:3])  # Últimos 3 días
        older_alerts = sum(row['cantidad'] for row in data[3:6])  # 3 días anteriores

        if recent_alerts > older_alerts * 1.2:
            direction = 'increasing'
        elif recent_alerts < older_alerts * 0.8:
            direction = 'decreasing'
        else:
            direction = 'stable'

        return {
            'trend': 'analyzed',
            'direction': direction,
            'recent_alerts': recent_alerts,
            'comparison_alerts': older_alerts
        }

    def create_alert_notification(self, alert_data, notification_type='email'):
        """Crea notificación de alerta"""
        try:
            notification = {
                'id': f"alert_{alert_data['patente']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                'type': notification_type,
                'priority': alert_data.get('alert_level', 'NORMAL'),
                'truck_data': alert_data,
                'message': self._generate_notification_message(alert_data),
                'created_at': datetime.now().isoformat(),
                'status': 'pending'
            }

            # En producción, aquí se enviaría la notificación real
            logger.info(f"Notificación creada: {notification['id']} para {alert_data['patente']}")

            return notification

        except Exception as e:
            logger.error(f"Error creando notificación: {e}")
            return None

    def _generate_notification_message(self, alert_data):
        """Genera mensaje de notificación"""
        patente = alert_data.get('patente', 'UNKNOWN')
        nivel = alert_data.get('alert_level', 'NORMAL')
        tiempo_espera = alert_data.get('tiempo_espera_horas', 0)
        destino = alert_data.get('deposito_destino', 'Destino desconocido')

        emoji_map = {
            'CRITICAL': '🚨',
            'WARNING': '⚠️',
            'ATTENTION': '🔔'
        }

        emoji = emoji_map.get(nivel, '📍')

        if nivel == 'CRITICAL':
            urgency = 'URGENTE'
            action = 'Requiere atención inmediata'
        elif nivel == 'WARNING':
            urgency = 'IMPORTANTE'
            action = 'Revisar en próximas 2 horas'
        else:
            urgency = 'INFORMACIÓN'
            action = 'Monitorear evolución'

        message = f"""
    {emoji} ALERTA {urgency} - CBN Tracking

    Camión: {patente}
    Destino: {destino}
    Tiempo de espera: {tiempo_espera:.1f} horas
    Estado: {alert_data.get('estado_entrega', 'EN_PROCESO')}

    {action}

    Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        """.strip()

        return message

    def __del__(self):
        """Destructor para limpiar conexiones automáticamente"""
        self.disconnect_databases()

